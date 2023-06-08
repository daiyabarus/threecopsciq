import csv
import traceback
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font

redFill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
darkRedText = Font(color="FF800000")
greenFill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
darkGreenText = Font(color="FF008000")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def to_csv(list_of_mo: list, file_to_save: str, csv_header: list) -> bool:
    with open(file_to_save, "w", encoding="UTF8", newline="") as f:
        writer = csv.writer(f)

        # write the header
        writer.writerow(csv_header)

        # write the data
        for item in list_of_mo:
            writer.writerow(item)

    return os.path.exists(file_to_save)


def to_txt(contents: str, file_to_save: str) -> bool:
    with open(file_to_save, "w", encoding="utf-8") as f:
        f.write(contents)

    return os.path.exists(file_to_save)


def to_txt_unix(contents: str, file_to_save: str) -> bool:
    with open(file_to_save, "w", newline="\n", encoding="utf-8") as f:
        f.write(contents)

    return os.path.exists(file_to_save)


def to_txt_unix_append(contents: str, file_to_save: str) -> bool:
    with open(file_to_save, "a", newline="\n", encoding="utf-8") as f:
        f.write(contents)

    return os.path.exists(file_to_save)


def to_xlsx(
    file_to_save: str, ws_name: str, list_of_header: list, list_of_contents: list
):
    # check if workbook exists
    if os.path.exists(file_to_save):
        wb = load_workbook(file_to_save)
    else:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    # create cell
    ws_target = wb.create_sheet(ws_name)

    # starting row
    starting_row = 1

    # header
    for index, header in enumerate(list_of_header):
        ws_target.cell(row=starting_row, column=index + 1).value = header

    starting_row += 1
    for items in list_of_contents:
        for col, item in enumerate(items):
            ws_target.cell(row=starting_row, column=col + 1).value = item

        starting_row += 1

    wb.save(file_to_save)

    return os.path.exists(file_to_save)


def to_xlsx_undefined_filled(
    file_to_save: str,
    ws_name: str,
    list_of_header: list,
    list_of_contents: list,
    list_of_red: list = None,
    list_of_green: list = None,
):
    # check if workbook exists
    if os.path.exists(file_to_save):
        wb = load_workbook(file_to_save)
    else:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    # create cell
    ws_target = wb.create_sheet(ws_name)
    starting_row = 1

    # header
    for index, header in enumerate(list_of_header):
        ws_target.cell(row=starting_row, column=index + 1).value = header

    starting_row += 1
    for items in list_of_contents:
        for col, item in enumerate(items):
            ws_target.cell(row=starting_row, column=col + 1).value = item
            if list_of_red and any(ext in item for ext in list_of_red):
                ws_target.cell(row=starting_row, column=col + 1).fill = redFill
                ws_target.cell(row=starting_row, column=col + 1).font = darkRedText
            elif list_of_green and any(ext in item for ext in list_of_green):
                ws_target.cell(row=starting_row, column=col + 1).fill = greenFill
                ws_target.cell(row=starting_row, column=col + 1).font = darkGreenText

        starting_row += 1

    wb.save(file_to_save)

    return os.path.exists(file_to_save)


def to_xlsx_filled_and_col_offside(
    file_to_save: str,
    ws_name: str,
    list_of_header: list,
    list_of_contents: list,
    list_of_red: list = None,
    list_of_green: list = None,
    col_offside: int = 0,
):
    try:
        # check if workbook exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create cell
        if ws_name in wb.sheetnames:
            ws_target = wb[ws_name]
        else:
            ws_target = wb.create_sheet(ws_name)

        # starting row
        starting_row = 1

        # header
        for index, header in enumerate(list_of_header):
            col = index + 1 + col_offside
            ws_target.cell(row=starting_row, column=col).value = header

        starting_row += 1
        for items in list_of_contents:
            for col, item in enumerate(items):
                ws_target.cell(
                    row=starting_row, column=col + 1 + col_offside
                ).value = item
                if list_of_red and any(ext in item for ext in list_of_red):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).fill = redFill
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = darkRedText
                elif list_of_green and any(ext in item for ext in list_of_green):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).fill = greenFill
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = darkGreenText

            starting_row += 1

        wb.save(file_to_save)

        return True

    except Exception:
        errors = traceback.format_exc()
        print(errors)
        return False


def to_xlsx_offside_noheader(
    file_to_save: str,
    ws_name: str,
    list_of_contents: list,
    list_of_red: list = None,
    list_of_green: list = None,
    col_offside: int = 0,
):
    try:
        # check if workbook exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create cell
        if ws_name in wb.sheetnames:
            ws_target = wb[ws_name]
        else:
            ws_target = wb.create_sheet(ws_name)

        starting_row = 2
        for items in list_of_contents:
            for col, item in enumerate(items):
                ws_target.cell(
                    row=starting_row, column=col + 1 + col_offside
                ).value = item
                if list_of_red and item in list_of_red:
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).fill = redFill
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = darkRedText
                elif list_of_green and item in list_of_green:
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).fill = greenFill
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = darkGreenText

                ws_target.cell(
                    row=starting_row, column=col + 1 + col_offside
                ).border = thin_border

                ws_target.cell(
                    row=starting_row, column=col + 1 + col_offside
                ).alignment = Alignment(horizontal="center", vertical="center")

                ws_target.cell(
                    row=starting_row, column=col + 1 + col_offside
                ).font = Font(name="Calibri", size=11)

            starting_row += 1

        wb.save(file_to_save)

        return True

    except Exception:
        errors = traceback.format_exc()
        print(errors)
        return False
