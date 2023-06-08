from openpyxl import load_workbook

from utils.to_get import ToGet


class ExcelReader:
    """
    Base Class for Excel Functions
    """

    def __init__(self, excel_path: str, excel_profile: str) -> None:
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, data_only=True, read_only=True)
        self.dtprofile = ToGet.jsonfile_to_dict(jsonfile=excel_profile)
        self.dtheader = self.dtprofile.get("header_row", {})

    def get_index_col(self, header_name: str, headerlist: tuple) -> int:
        try:
            return headerlist.index(header_name)
        except ValueError as exc:
            raise ValueError(f"Header Name: {header_name} not found!") from exc

    def get_headerlist(
        self,
        wsname: str,
        header_row: int,
    ) -> tuple:
        try:
            ws = self.wb[wsname]

            for row in ws.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ):
                return tuple(row)
        except Exception as exc:
            raise type(exc)(f"Failed to get header list: {exec}") from exc
        return tuple()

    def get_sheetnames(self):
        return self.wb.sheetnames

    def is_sheet_in(self, sheetname: str) -> bool:
        return sheetname in self.wb.sheetnames
